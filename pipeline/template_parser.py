"""
Extracts structured data from raw text using a YAML customer template.

Each template defines regex patterns for each field. The parser tries
each pattern and returns the first match. Fields with no match return None.
"""
import logging
import re
from pathlib import Path
from typing import Optional

import yaml

from config.settings import TEMPLATES_DIR

log = logging.getLogger(__name__)


def _load_template(template_name: str) -> Optional[dict]:
    template_path = TEMPLATES_DIR / f"{template_name}.yaml"
    if not template_path.exists():
        return None
    try:
        return yaml.safe_load(template_path.read_text(encoding="utf-8"))
    except Exception as e:
        log.warning(f"Failed to load template {template_name}.yaml: {e}")
        return None


def _extract_field(text: str, patterns: list[str]) -> Optional[str]:
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
        if match:
            # Join all capture groups (supports multi-group date patterns); fall back to full match
            return (" ".join(g for g in match.groups() if g) if match.lastindex else match.group(0)).strip()
    return None


def _extract_line_items(text: str, pattern_or_patterns) -> list[dict]:
    """
    Extracts line items using one or more regex patterns with named groups.
    Supports both a single pattern string and a list of patterns (line_items_patterns).
    Each pattern should use named groups: product_code, description, qty, uom,
    unit_price, subtotal, total — any subset is fine.
    Results from all patterns are merged; duplicate rows are deduplicated.
    """
    patterns = (
        pattern_or_patterns
        if isinstance(pattern_or_patterns, list)
        else [pattern_or_patterns]
    )
    seen: set = set()
    items: list[dict] = []
    for pattern in patterns:
        if not pattern:
            continue
        for match in re.finditer(pattern, text, re.IGNORECASE | re.MULTILINE):
            item = {k: v.strip() if v else None for k, v in match.groupdict().items()}
            key = frozenset(item.items())
            if key not in seen:
                seen.add(key)
                items.append(item)
    return items


def _find_cell_value(ws, label: str):
    """
    Search all cells in a worksheet for a cell whose string value matches label
    (case-insensitive). Returns the value of the adjacent cell to the right,
    or the cell directly below, whichever is non-empty. Returns None if not found.
    """
    label_lower = label.lower().strip()
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and str(cell.value).lower().strip() == label_lower:
                # Try right neighbour first
                right = ws.cell(row=cell.row, column=cell.column + 1)
                if right.value not in (None, ""):
                    return str(right.value).strip()
                # Try cell below
                below = ws.cell(row=cell.row + 1, column=cell.column)
                if below.value not in (None, ""):
                    return str(below.value).strip()
    return None


def _extract_xlsx_line_items(wb, config: dict) -> list[dict]:
    """
    Extract line items from an Excel workbook using the `line_items_xlsx` config block.

    Config keys:
      sheet         — sheet index (int, 0-based) or name (str). Default: 0.
      header_row    — 1-indexed row containing column headers. Default: 1.
      columns       — dict mapping named-group → column header string.
      skip_if_empty — named-group key; skip row if that column is empty.
    """
    sheet_ref = config.get("sheet", 0)
    if isinstance(sheet_ref, int):
        ws = wb.worksheets[sheet_ref]
    else:
        ws = wb[sheet_ref]

    header_row_idx = config.get("header_row", 1)  # 1-indexed
    col_map: dict[str, str] = config.get("columns", {})  # named_group → header_label
    skip_key: str = config.get("skip_if_empty", "")

    # Build column_index → named_group from header row
    header_cells = list(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))[0]
    col_index: dict[str, int] = {}  # named_group → 0-based col index
    for named_group, label in col_map.items():
        label_lower = label.lower().strip()
        for idx, cell_val in enumerate(header_cells):
            if cell_val and str(cell_val).lower().strip() == label_lower:
                col_index[named_group] = idx
                break

    if not col_index:
        return []

    items: list[dict] = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        item = {}
        for named_group, idx in col_index.items():
            val = row[idx] if idx < len(row) else None
            item[named_group] = str(val).strip() if val is not None else None
        if skip_key and not item.get(skip_key):
            continue
        if any(v for v in item.values()):
            items.append(item)

    return items


def parse_xlsx(path: Path, template_name: str) -> dict:
    """
    Parse an Excel workbook using a template's `fields_xlsx` and `line_items_xlsx` sections.
    Falls back to regex on flat text for fields not covered by `fields_xlsx`.
    Returns same shape as parse(): dict with field values + _confidence.
    """
    import openpyxl
    tmpl = _load_template(template_name)
    if not tmpl:
        log.warning(f"No template found: {template_name!r}")
        return {"error": "no_template", "_confidence": 0.0}

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]

    fields_xlsx: dict = tmpl.get("fields_xlsx", {})  # field_name → label string
    fields_regex: dict = tmpl.get("fields", {})       # field_name → [patterns]
    required_fields: list = tmpl.get("required_fields", list(fields_regex.keys()))

    result: dict = {}

    # 1. Cell-label lookup for fields_xlsx
    for field_name, label in fields_xlsx.items():
        result[field_name] = _find_cell_value(ws, label)

    # 2. Flat-text regex for any remaining fields not already found
    if fields_regex:
        from pipeline.text_extractor import extract_excel_text
        flat_text = extract_excel_text(path)
        for field_name, patterns in fields_regex.items():
            if field_name not in result or result[field_name] is None:
                result[field_name] = _extract_field(
                    flat_text, patterns if isinstance(patterns, list) else [patterns]
                )

    # 3. Line items from line_items_xlsx config
    xlsx_li_config = tmpl.get("line_items_xlsx")
    if xlsx_li_config:
        result["line_items"] = _extract_xlsx_line_items(wb, xlsx_li_config)
    else:
        result["line_items"] = []

    extracted = sum(1 for f in required_fields if result.get(f))
    total_slots = len(required_fields)
    min_li = tmpl.get("min_line_items", 0)
    if min_li:
        total_slots += 1
        if len(result.get("line_items", [])) >= min_li:
            extracted += 1
    confidence = extracted / total_slots if total_slots else 0.0
    result["_confidence"] = round(confidence, 2)
    result["_required_fields_matched"] = extracted
    result["_required_fields_total"] = total_slots

    log.info(f"Parsed xlsx {template_name!r}: confidence={confidence:.0%}, "
             f"{extracted}/{total_slots} required slots filled.")
    return result


def parse(text: str, template_name: str) -> dict:
    """
    Returns a dict of extracted fields. Missing fields are None.
    Confidence is a simple ratio of non-null required fields.
    template_name is the YAML file stem (e.g. "evergy", not "evergy.yaml").
    """
    tmpl = _load_template(template_name)
    if not tmpl:
        log.warning(f"No template found: {template_name!r}")
        return {"error": "no_template", "_confidence": 0.0}

    fields = tmpl.get("fields", {})
    required_fields = tmpl.get("required_fields", list(fields.keys()))

    result = {}
    for field_name, patterns in fields.items():
        result[field_name] = _extract_field(text, patterns if isinstance(patterns, list) else [patterns])

    # Support both line_items_patterns (list) and legacy line_items_pattern (string)
    line_patterns = tmpl.get("line_items_patterns") or tmpl.get("line_items_pattern", "")
    result["line_items"] = _extract_line_items(text, line_patterns)

    # Confidence = fraction of required fields that were extracted.
    # min_line_items (if set) counts as one additional required slot.
    extracted = sum(1 for f in required_fields if result.get(f))
    total_slots = len(required_fields)
    min_li = tmpl.get("min_line_items", 0)
    if min_li:
        total_slots += 1
        if len(result["line_items"]) >= min_li:
            extracted += 1
    confidence = extracted / total_slots if total_slots else 0.0
    result["_confidence"] = round(confidence, 2)
    result["_required_fields_matched"] = extracted
    result["_required_fields_total"] = total_slots

    log.info(f"Parsed {template_name!r}: confidence={confidence:.0%}, "
             f"{extracted}/{total_slots} required slots filled.")

    # Record stat for trend analysis — best-effort, never fatal
    try:
        from database import db
        db.record_template_stat(
            template_name=template_name,
            confidence=round(confidence, 4),
            required_fields_matched=extracted,
            required_fields_total=len(required_fields),
        )
    except Exception:
        pass

    return result
